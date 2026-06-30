from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _row_text(row: tuple[Any, ...]) -> str:
    cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
    return " | ".join(cells)


def format_workbook_text(workbook_path: Path) -> str:
    """Convert客服学习手册 xlsx into plain text for LLM system context."""
    wb = load_workbook(workbook_path, data_only=True)
    sections: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines: list[str] = []
        for row in ws.iter_rows(values_only=True):
            line = _row_text(tuple(row))
            if line:
                lines.append(line)
        if lines:
            sections.append(f"### {sheet_name}\n" + "\n".join(lines))

    wb.close()
    return "\n\n".join(sections).strip()


def load_knowledge_text(path: Path | None) -> str:
    if not path:
        return ""

    resolved = path.expanduser()
    if not resolved.is_absolute():
        resolved = Path.cwd() / resolved
    if not resolved.exists():
        return ""

    suffix = resolved.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return format_workbook_text(resolved)
    if suffix in {".txt", ".md"}:
        return resolved.read_text(encoding="utf-8").strip()
    return ""


def append_reference_section(
    prompt: str,
    title: str,
    intro: str,
    body: str,
    *,
    max_chars: int = 0,
) -> str:
    text = str(body or "").strip()
    if not text:
        return str(prompt or "").strip()
    if max_chars and len(text) > max_chars:
        text = text[:max_chars] + "\n…（内容过长，已截断）"
    block = f"## {title}\n{intro}\n\n{text}"
    base = str(prompt or "").strip()
    return f"{base}\n\n{block}" if base else block


def build_system_prompt(base_prompt: str, knowledge_text: str, *, max_chars: int = 28000) -> str:
    base = str(base_prompt or "").strip()
    knowledge = str(knowledge_text or "").strip()
    if not knowledge:
        return base

    if len(knowledge) > max_chars:
        knowledge = knowledge[:max_chars] + "\n…（手册内容过长，已截断）"

    handbook_block = (
        "## 和田玉手镯客服学习手册（必须遵守）\n"
        "以下内容为店铺官方客服手册，回复买家时必须优先遵循手册中的标准话术、口径与禁止事项。"
        "不得与手册矛盾，不得编造订单、物流、赔付等未在手册中授权的承诺。\n\n"
        f"{knowledge}"
    )

    if base:
        return f"{base}\n\n{handbook_block}"
    return handbook_block
